from selenium import webdriver
from selenium.webdriver.support.select import Select 
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import time

class usuarioAcessorias:
    def __init__(self, driver, user, password, spreadsheet):
        self.usuario = user
        self.senha = password
        self.planilha = spreadsheet
        self.driver = driver

    def fazerLogin(self):
        driver = self.driver
        campo_user = driver.find_element(By.XPATH, "//input[@name='mailAC']")
        campo_user.clear()
        campo_user.send_keys(self.usuario)
        time.sleep(2)

        campo_senha = driver.find_element(By.XPATH, "//input[@name='passAC']")
        campo_senha.clear()
        campo_senha.send_keys(self.senha)
        campo_senha.send_keys(Keys.RETURN)
        time.sleep(2)

    def carregarCNPJ(self, planilha):
        wb = openpyxl.load_workbook(planilha)
        ws = wb.active

        cnpjs = []
        try:
            if ws.cell(row=1, column=1).value == 'CNPJ':
                for row in ws.iter_rows(min_row=2, values_only=True):
                    cnpj = row[0]  # Supondo que a primeira coluna contenha os CNPJs
                    cnpjs.append(cnpj)
            else:
                mb.showerror(title ="Erro", message = "Nenhuma empresa encontrada na planilha")

        except:            
            mb.showerror(title ="Erro", message = "Nenhuma empresa encontrada na planilha")

        return cnpjs

    def marcarDepartamento(self):
        time.sleep(1)
        driver = self.driver

        botaoMenu = driver.find_element(By.XPATH, "//*[@id='menu-toggler']")
        botaoMenu.click()

        time.sleep(2)

        # Clicar na aba "Empresa"
        campo_empresa = driver.find_element(By.XPATH, "//a[@href='sysmain.php?m=4']")
        campo_empresa.click()
        time.sleep(2)
        
        empresas = self.carregarCNPJ(self.planilha)
        for cnpj in empresas:
            time.sleep(1)
            campo_busca = driver.find_element(By.XPATH, "//input[@name='searchString']")
            campo_busca.clear()
            campo_busca.send_keys(cnpj)
            campo_busca.send_keys(Keys.RETURN)
            time.sleep(2)

            try:
                # Clicar no primeiro resultado de empresa encontrado
                campo_empresaId = driver.find_element(By.XPATH, "//*[@id='divEmpresas']/div[1]")
                campo_empresaId.click()
                time.sleep(2)
                driver.execute_script("window.scrollTo(0, 300)")



                # Listas de XPath para as interações
                clicks1 = ["//*[@id='CttEdit_0_1']/button[1]", "//*[@id='CttEdit_0_2']/button[1]", "//*[@id='CttEdit_0_3']/button[1]", "//*[@id='CttEdit_0_4']/button[1]"]
                clicks2 = ["//*[@id='divCtt_0_1']/div[1]/div/span[1]/button", "//*[@id='divCtt_0_2']/div[1]/div/span[1]/button", "//*[@id='divCtt_0_3']/div[1]/div/span[1]/button", "//*[@id='divCtt_0_4']/div[1]/div/span[1]/button"]
                clicks3 = ["//*[@id='selDpZ_0_1']/a[1]", "//*[@id='selDpZ_0_2']/a[1]", "//*[@id='selDpZ_0_3']/a[1]", "//*[@id='selDpZ_0_4']/a[1]"]
                clicks4 = ["//*[@id='CttSave_0_1']/button[1]", "//*[@id='CttSave_0_2']/button[1]", "//*[@id='CttSave_0_3']/button[1]", "//*[@id='CttSave_0_4']/button[1]"]

                # Clica no ícone para editar o contato
                for x in clicks1:
                    try:
                        click01 = driver.find_element(By.XPATH, x)
                        click01.click()
                        time.sleep(1)
                    except:
                        pass

                # Clica no chuveirinho para abrir os departamentos
                for y in clicks2:
                    try:
                        click02 = driver.find_element(By.XPATH, y)
                        click02.click()
                        time.sleep(1)
                    except:
                        pass

                # Marca todos os departamentos dos contatos
                for z in clicks3:
                    try:
                        click03 = driver.find_element(By.XPATH, z)
                        click03.click()
                        time.sleep(1)
                    except:
                        pass

                # Clica em salvar os dados
                for s in clicks4:
                    try:
                        click04 = driver.find_element(By.XPATH, s)
                        click04.click()
                        time.sleep(3)
                    except:
                        pass

                # Clica em voltar para o campo de busca
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                campo_voltar = driver.find_element(By.XPATH, '//*[@id="navFim"]/div[2]/button[2]')
                campo_voltar.click()
                time.sleep(4)

                print("Contatos marcados para o CNPJ:", cnpj)

            except:
                print("Erro ao interagir com o CNPJ:", cnpj)



    def marcarEnvioAgendado(self): # FEATURE FLAG
        time.sleep(1)
        driver = self.driver
        econtinuo = driver.find_element(By.XPATH, "//*[@id='M1']/a/b")

        botaoMenu = driver.find_element(By.XPATH, "//*[@id='menu-toggler']")
        botaoMenu.click()

        time.sleep(1)

        ActionChains(driver)\
        .move_to_element(econtinuo)\
        .perform()
        time.sleep(2)

        econtinuoUm = driver.find_element(By.XPATH, "//*[@id='M130']/a")
        ActionChains(driver)\
        .move_to_element(econtinuoUm)\
        .perform()
        time.sleep(2)

        econtinuoDois = driver.find_element(By.XPATH, "//*[@id='M128']")
        ActionChains(driver)\
        .move_to_element(econtinuoDois)\
        .perform()
        econtinuoDois.click()
        time.sleep(3)



        while True:
                editarPagina(self)
                time.sleep(1)
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                proximaPagina = driver.find_element(By.XPATH, '//*[@id="main-container"]/div[2]/div[2]/div/div/div[15]/ul/li[9]/a/i')
                time.sleep(1)
                
                try:
                    proximaPagina.click()
                    
                except:
                    print('robôzinho finalizou pros cria 8D')

def editarPagina(self):
    driver = self.driver
    listaEditar = ['//*[@id="main-container"]/div[2]/div[2]/div/div/div[3]/div[1]/span/a[2]/i', '//*[@id="main-container"]/div[2]/div[2]/div/div/div[4]/div[1]/span/a[2]/i', '//*[@id="main-container"]/div[2]/div[2]/div/div/div[5]/div[1]/span/a[2]/i', '//*[@id="main-container"]/div[2]/div[2]/div/div/div[6]/div[1]/span/a[2]/i', '//*[@id="main-container"]/div[2]/div[2]/div/div/div[7]/div[1]/span/a[2]/i', '//*[@id="main-container"]/div[2]/div[2]/div/div/div[8]/div[1]/span/a[2]/i', '//*[@id="main-container"]/div[2]/div[2]/div/div/div[9]/div[1]/span/a[2]/i', '//*[@id="main-container"]/div[2]/div[2]/div/div/div[10]/div[1]/span/a[2]/i', '//*[@id="main-container"]/div[2]/div[2]/div/div/div[11]/div[1]/span/a[2]/i', '//*[@id="main-container"]/div[2]/div[2]/div/div/div[12]/div[1]/span/a[2]/i', '//*[@id="main-container"]/div[2]/div[2]/div/div/div[13]/div[1]/span/a[2]/i', '//*[@id="main-container"]/div[2]/div[2]/div/div/div[14]/div[1]/span/a[2]/i']

    for botaoEditar in listaEditar:      
        editar = driver.find_element(By.XPATH, botaoEditar)
        editar.click()
        time.sleep(2)

        agendado = "Sim - Agendado"
        emailimediato = driver.find_element(By.XPATH,"//*[@id='EclSendMail']")
        emailimediato.send_keys(agendado)
        emailimediato.send_keys(Keys.RETURN)
        time.sleep(3)

        botaoSalvar = driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div[2]/div/div/form/div[4]/div[3]/div/button[1]')
        botaoSalvar.click()
        time.sleep(3)




